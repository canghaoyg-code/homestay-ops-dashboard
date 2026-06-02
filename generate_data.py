from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "酒店民宿运营SOP数据看板.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "dashboard-data.json"


def value(v):
    return "" if v is None else v


def rows(ws, start=5, max_row=40):
    header = [value(c.value) for c in ws[4]]
    items = []
    for row in ws.iter_rows(min_row=start, max_row=max_row, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        item = {str(header[i]): value(row[i]) for i in range(min(len(header), len(row)))}
        items.append(item)
    return items


def count(items, predicate):
    return sum(1 for item in items if predicate(item))


def num(v):
    try:
        if v in ("", None):
            return 0
        return float(v)
    except Exception:
        return 0


def avg(values):
    clean = [num(v) for v in values if num(v) != 0]
    return sum(clean) / len(clean) if clean else 0


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    wb = load_workbook(xlsx_path, data_only=True)

    leads = rows(wb["01线索登记"])
    customers = rows(wb["02客户分级"])
    properties = rows(wb["03房源评估"])
    quotes = rows(wb["04报价测算"])
    repairs = rows(wb["09维修记录"])
    owner_reports = rows(wb["10房东月报"])
    reviews = rows(wb["11周月复盘"])

    valid_leads = [item for item in leads if item.get("线索ID") and item.get("当前状态") != "无效"]
    a_customers = [item for item in customers if item.get("客户分级") == "A"]
    recommended_properties = [item for item in properties if item.get("结论") == "建议接"]
    pending_repairs = [item for item in repairs if item.get("维修ID") and item.get("处理状态") != "已完成"]

    source_labels = ["抖音", "微信", "电话", "朋友转介绍", "其他"]
    property_labels = ["建议接", "谨慎接", "不建议接", "需补充信息"]

    estimated_revenue = sum(num(item.get("月收入")) for item in quotes)
    estimated_net = sum(num(item.get("月净收益")) for item in quotes)
    monthly_revenue = sum(num(item.get("总收入")) for item in owner_reports)
    monthly_net = sum(num(item.get("净收益")) for item in owner_reports)
    avg_assessment = avg([item.get("总分") for item in properties])
    avg_payback = avg([item.get("预计回本月数") for item in quotes])
    avg_occupancy = avg([item.get("入住率") for item in owner_reports])

    data = {
        "meta": {
            "title": "酒店 / 民宿运营数据看板",
            "source": xlsx_path.name,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
        },
        "kpis": [
            {"label": "新增线索", "value": len([x for x in leads if x.get("线索ID")]), "unit": "条", "note": "线索池规模"},
            {"label": "有效线索", "value": len(valid_leads), "unit": "条", "note": "需持续跟进"},
            {"label": "A级客户", "value": len(a_customers), "unit": "位", "note": "当天优先联系"},
            {"label": "建议接房源", "value": len(recommended_properties), "unit": "套", "note": "可进入报价"},
            {"label": "本月净收益", "value": monthly_net, "unit": "元", "note": "房东月报口径", "format": "money"},
        ],
        "charts": {
            "funnel": [
                {"label": "新增线索", "value": len([x for x in leads if x.get("线索ID")])},
                {"label": "有效线索", "value": len(valid_leads)},
                {"label": "A级客户", "value": len(a_customers)},
                {"label": "预约看房", "value": count(leads, lambda x: x.get("当前状态") == "已约看房")},
                {"label": "已签约", "value": count(leads, lambda x: x.get("当前状态") == "已签约")},
            ],
            "sources": [{"label": label, "value": count(leads, lambda x, label=label: x.get("来源平台") == label)} for label in source_labels],
            "properties": [{"label": label, "value": count(properties, lambda x, label=label: x.get("结论") == label)} for label in property_labels],
            "revenue": [
                {"label": "预估月收入", "value": estimated_revenue},
                {"label": "预估月净收益", "value": estimated_net},
                {"label": "本月总收入", "value": monthly_revenue},
                {"label": "本月净收益", "value": monthly_net},
            ],
        },
        "risks": [
            {
                "module": "线索",
                "metric": "待初筛/待资料",
                "value": count(leads, lambda x: x.get("当前状态") in ("待初筛", "待资料")),
                "rule": "当天状态仍未推进",
                "action": "当天清理，避免线索过期",
            },
            {
                "module": "房源",
                "metric": "平均评估分",
                "value": round(avg_assessment, 1),
                "rule": "低于20分谨慎推进",
                "action": "复核物业、竞品和投入风险",
            },
            {
                "module": "收益",
                "metric": "平均回本月数",
                "value": round(avg_payback, 1),
                "rule": "超过24个月需重新评估",
                "action": "调整投入、定价或放弃接房",
            },
            {
                "module": "维修",
                "metric": "未完成维修",
                "value": len(pending_repairs),
                "rule": "影响入住问题优先",
                "action": "优先处理高优先级维修",
            },
            {
                "module": "房东",
                "metric": "平均入住率",
                "value": round(avg_occupancy * 100, 1),
                "unit": "%",
                "rule": "低入住率房源进入复盘",
                "action": "优化首图、价格和入住说明",
            },
        ],
        "tables": {
            "hotLeads": [
                {
                    "id": item.get("线索ID"),
                    "name": item.get("客户称呼"),
                    "area": item.get("房源区域"),
                    "need": item.get("需求类型"),
                    "status": item.get("当前状态"),
                    "next": item.get("下次跟进"),
                }
                for item in leads[:8]
                if item.get("线索ID")
            ],
            "repairs": [
                {
                    "id": item.get("维修ID"),
                    "property": item.get("房源ID"),
                    "type": item.get("问题类型"),
                    "priority": item.get("优先级"),
                    "status": item.get("处理状态"),
                    "due": item.get("预计完成"),
                }
                for item in repairs[:8]
                if item.get("维修ID")
            ],
            "reviews": [
                {
                    "period": item.get("周期"),
                    "leads": item.get("新增线索"),
                    "valid": item.get("有效线索"),
                    "signed": item.get("签约数"),
                    "issue": item.get("重复问题"),
                    "action": item.get("下步动作"),
                }
                for item in reviews[:6]
                if item.get("周期")
            ],
        },
    }

    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
